"""
Excel Exporter for Evaluation and Comparison Results.

Generates .xlsx files with a multi-sheet schema optimised for Power BI
consumption (star-schema, snake_case columns, long/tidy metrics table).
"""

from __future__ import annotations

import io
import logging
from datetime import datetime
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, numbers
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Metric-unit mapping (for the long Metrics table)
# ---------------------------------------------------------------------------
_METRIC_UNITS: Dict[str, str] = {
    # classification
    "accuracy": "ratio", "precision": "ratio", "recall": "ratio",
    "f1_score": "ratio", "kappa": "ratio",
    "subcategory_accuracy": "ratio", "priority_accuracy": "ratio",
    "sentiment_accuracy": "ratio", "avg_confidence": "ratio",
    # quality
    "relevance": "ratio", "groundedness": "ratio",
    "format_compliance": "ratio", "completeness": "ratio",
    "instruction_following": "ratio", "empathy_score": "ratio",
    "follow_up_quality": "ratio", "rule_compliance": "ratio",
    "optimal_similarity": "ratio", "resolution_efficiency": "ratio",
    "question_count_avg": "count",
    # latency
    "mean_latency": "seconds", "median_latency": "seconds",
    "p95_latency": "seconds", "p99_latency": "seconds",
    "min_latency": "seconds", "max_latency": "seconds",
    "std_latency": "seconds", "mean_ttft": "seconds",
    "tokens_per_second": "tokens/s",
    "cost_per_request": "USD", "total_cost": "USD",
    "cache_hit_rate": "%", "reasoning_token_pct": "%",
    "avg_prompt_tokens": "tokens", "avg_completion_tokens": "tokens",
    # consistency
    "reproducibility_score": "ratio", "semantic_similarity": "ratio",
    "format_consistency": "ratio", "variance_coefficient": "ratio",
    # tool calling
    "tool_selection_accuracy": "ratio", "parameter_accuracy": "ratio",
    "execution_success_rate": "ratio",
}

# Header styling
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _auto_width(ws, min_width: int = 10, max_width: int = 50) -> None:
    """Auto-fit column widths based on content."""
    for col_cells in ws.columns:
        lengths = []
        for cell in col_cells:
            val = str(cell.value) if cell.value is not None else ""
            lengths.append(min(len(val), max_width))
        best = max(lengths) + 2 if lengths else min_width
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = max(best, min_width)


def _write_header(ws, headers: List[str]) -> None:
    """Write a styled header row."""
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGNMENT
    ws.freeze_panes = "A2"


def _safe_value(val: Any) -> Any:
    """Convert numpy/unusual types to native Python for openpyxl."""
    if val is None:
        return None
    try:
        import numpy as np
        if isinstance(val, (np.integer,)):
            return int(val)
        if isinstance(val, (np.floating,)):
            return float(val)
        if isinstance(val, np.ndarray):
            return val.tolist()
    except ImportError:
        pass
    if isinstance(val, (list, dict)):
        import json
        return json.dumps(val, ensure_ascii=False, default=str)
    return val


def _flatten_metric_dict(
    metrics: Optional[Dict[str, Any]], category: str, model_name: str, eval_type: str
) -> List[Dict[str, Any]]:
    """Flatten a metrics dict into long-format rows for the Metrics sheet."""
    if not metrics:
        return []
    rows = []
    for key, val in metrics.items():
        if val is None:
            continue
        # Skip nested dicts (e.g. category_accuracy, confusion_matrix)
        if isinstance(val, (dict, list)):
            continue
        try:
            numeric = float(val)
        except (TypeError, ValueError):
            continue
        rows.append({
            "model_name": model_name,
            "evaluation_type": eval_type,
            "metric_category": category,
            "metric_name": key,
            "metric_value": numeric,
            "metric_unit": _METRIC_UNITS.get(key, ""),
        })
    return rows


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

class ExcelExporter:
    """Generates Power BI–ready .xlsx workbooks from result JSON data."""

    # -----------------------------------------------------------------------
    # Evaluation export
    # -----------------------------------------------------------------------
    @staticmethod
    def export_evaluation(data: Dict[str, Any]) -> io.BytesIO:
        """Export a single evaluation result to an Excel workbook.

        Sheets: Metadata, Metrics, RawResults, CategoryAccuracy*, FoundryScores*
        (* = only when data is available)
        """
        wb = Workbook()
        model = data.get("model_name", "unknown")
        eval_type = data.get("evaluation_type", "unknown")

        # --- Sheet 1: Metadata ---
        ws_meta = wb.active
        ws_meta.title = "Metadata"
        _write_header(ws_meta, [
            "export_timestamp", "model_name", "evaluation_type",
            "eval_timestamp", "scenarios_tested", "error_count", "errors",
        ])
        ws_meta.append([
            datetime.utcnow().isoformat(),
            model,
            eval_type,
            data.get("timestamp", ""),
            data.get("scenarios_tested", 0),
            data.get("error_count", len(data.get("errors", []))),
            "; ".join(data.get("errors", [])) or None,
        ])
        _auto_width(ws_meta)

        # --- Sheet 2: Metrics (long / tidy) ---
        ws_metrics = wb.create_sheet("Metrics")
        _write_header(ws_metrics, [
            "model_name", "evaluation_type", "metric_category",
            "metric_name", "metric_value", "metric_unit",
        ])
        metric_rows: List[Dict[str, Any]] = []
        for category, key in [
            ("classification", "classification_metrics"),
            ("quality", "quality_metrics"),
            ("latency", "latency_metrics"),
            ("consistency", "consistency_metrics"),
            ("tool_calling", "tool_calling_metrics"),
            ("realtime", "realtime_metrics"),
        ]:
            metric_rows.extend(
                _flatten_metric_dict(data.get(key), category, model, eval_type)
            )
        for row in metric_rows:
            ws_metrics.append([
                row["model_name"], row["evaluation_type"], row["metric_category"],
                row["metric_name"], row["metric_value"], row["metric_unit"],
            ])
        _auto_width(ws_metrics)

        # --- Sheet 3: RawResults (wide, union schema) ---
        raw = data.get("raw_results") or []
        if raw:
            ws_raw = wb.create_sheet("RawResults")
            ExcelExporter._write_raw_results(ws_raw, raw, model, eval_type)

        # --- Sheet 4: CategoryAccuracy (optional) ---
        cat_acc = (data.get("classification_metrics") or {}).get("category_accuracy")
        if cat_acc and isinstance(cat_acc, dict):
            ws_cat = wb.create_sheet("CategoryAccuracy")
            _write_header(ws_cat, ["model_name", "evaluation_type", "category", "accuracy"])
            for cat, acc in sorted(cat_acc.items()):
                ws_cat.append([model, eval_type, cat, _safe_value(acc)])
            _auto_width(ws_cat)

        # --- Sheet 5: FoundryScores (optional) ---
        foundry = data.get("foundry_scores")
        if foundry:
            ws_foundry = wb.create_sheet("FoundryScores")
            ExcelExporter._write_foundry_scores_eval(
                ws_foundry, foundry, data, model, eval_type
            )

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    # -----------------------------------------------------------------------
    # Comparison export
    # -----------------------------------------------------------------------
    @staticmethod
    def export_comparison(data: Dict[str, Any]) -> io.BytesIO:
        """Export a comparison report to an Excel workbook.

        Sheets: Metadata, Dimensions, Recommendations,
                StatisticalSignificance*, MigrationReadiness*,
                FoundryComparison*
        """
        wb = Workbook()
        model_a = data.get("model_a", "model_a")
        model_b = data.get("model_b", "model_b")
        eval_type = data.get("evaluation_type", "unknown")
        summary = data.get("summary") or {}

        # --- Sheet 1: Metadata ---
        ws_meta = wb.active
        ws_meta.title = "Metadata"
        _write_header(ws_meta, [
            "export_timestamp", "model_a", "model_b", "evaluation_type",
            "comparison_timestamp", "overall_winner",
            "total_dimensions", "model_a_wins", "model_b_wins", "ties",
            "high_impact_dimensions", "batch_id",
        ])
        # Try different key patterns for wins
        a_wins = summary.get(f"{model_a}_wins", summary.get("model_a_wins", 0))
        b_wins = summary.get(f"{model_b}_wins", summary.get("model_b_wins", 0))
        ws_meta.append([
            datetime.utcnow().isoformat(),
            model_a,
            model_b,
            eval_type,
            data.get("timestamp", ""),
            summary.get("overall_winner", ""),
            summary.get("total_dimensions", 0),
            a_wins,
            b_wins,
            summary.get("ties", 0),
            "; ".join(summary.get("high_impact_dimensions", [])),
            data.get("batch_id"),
        ])
        _auto_width(ws_meta)

        # --- Sheet 2: Dimensions ---
        ws_dim = wb.create_sheet("Dimensions")
        _write_header(ws_dim, [
            "model_a", "model_b", "evaluation_type", "dimension",
            "model_a_value", "model_b_value", "difference",
            "percent_change", "better_model", "significance",
        ])
        for d in data.get("dimensions") or []:
            ws_dim.append([
                model_a, model_b, eval_type,
                d.get("dimension", ""),
                _safe_value(d.get("model_a_value")),
                _safe_value(d.get("model_b_value")),
                _safe_value(d.get("difference")),
                _safe_value(d.get("percent_change")),
                d.get("better_model", ""),
                d.get("significance", ""),
            ])
        _auto_width(ws_dim)

        # --- Sheet 3: Recommendations ---
        recs = data.get("recommendations") or []
        if recs:
            ws_rec = wb.create_sheet("Recommendations")
            _write_header(ws_rec, [
                "model_a", "model_b", "evaluation_type",
                "recommendation_index", "recommendation_text",
            ])
            for idx, text in enumerate(recs, 1):
                ws_rec.append([model_a, model_b, eval_type, idx, text.strip()])
            _auto_width(ws_rec)

        # --- Sheet 4: StatisticalSignificance ---
        stats = data.get("statistical_significance")
        if stats:
            ws_stat = wb.create_sheet("StatisticalSignificance")
            _write_header(ws_stat, [
                "model_a", "model_b", "evaluation_type",
                "test_name", "statistic", "p_value", "significant",
                "detail_key", "detail_value",
            ])
            for test_name, test_data in stats.items():
                if not isinstance(test_data, dict):
                    continue
                base_row = [model_a, model_b, eval_type, test_name]
                stat_val = test_data.get("chi2", test_data.get("t_statistic"))
                p_val = test_data.get("p_value")
                sig = test_data.get("significant")
                # Main row
                ws_stat.append(base_row + [
                    _safe_value(stat_val), _safe_value(p_val), sig, None, None,
                ])
                # Detail rows (extra keys)
                skip = {"chi2", "t_statistic", "p_value", "significant"}
                for k, v in test_data.items():
                    if k not in skip:
                        ws_stat.append(base_row + [
                            None, None, None, k, _safe_value(v),
                        ])
            _auto_width(ws_stat)

        # --- Sheet 5: MigrationReadiness ---
        migration = data.get("migration_readiness")
        if migration and migration.get("verdict") != "NOT_CONFIGURED":
            ws_mig = wb.create_sheet("MigrationReadiness")
            _write_header(ws_mig, [
                "model_a", "model_b", "evaluation_type",
                "verdict", "target_model", "metric",
                "threshold", "actual", "passed",
            ])
            for check in migration.get("checks") or []:
                ws_mig.append([
                    model_a, model_b, eval_type,
                    migration.get("verdict", ""),
                    migration.get("model", ""),
                    check.get("metric", ""),
                    _safe_value(check.get("threshold")),
                    _safe_value(check.get("actual")),
                    check.get("passed"),
                ])
            _auto_width(ws_mig)

        # --- Sheet 6: FoundryComparison ---
        f_a = data.get("foundry_scores_a")
        f_b = data.get("foundry_scores_b")
        f_meta = data.get("foundry_meta") or {}
        if f_a or f_b:
            ws_fc = wb.create_sheet("FoundryComparison")
            _write_header(ws_fc, [
                "model", "evaluation_type", "grader_name",
                "aggregated_score", "report_url",
            ])
            for label, scores, meta_key in [
                (model_a, f_a, "model_a"),
                (model_b, f_b, "model_b"),
            ]:
                agg = (scores or {}).get("aggregated") or {}
                url = (f_meta.get(meta_key) or {}).get("report_url", "")
                for grader, score in agg.items():
                    ws_fc.append([
                        label, eval_type, grader,
                        _safe_value(score), url,
                    ])
            _auto_width(ws_fc)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    # -----------------------------------------------------------------------
    # Internal helpers
    # -----------------------------------------------------------------------
    @staticmethod
    def _write_raw_results(
        ws, raw: List[Dict[str, Any]], model: str, eval_type: str
    ) -> None:
        """Write the RawResults sheet with a union schema across all task types."""
        # Define all columns (common + per-type)
        columns = [
            # common
            "model_name", "evaluation_type", "scenario_id",
            "latency_s", "total_tokens",
            "prompt_tokens", "completion_tokens", "cached_tokens", "reasoning_tokens",
            # classification
            "input_text",
            "expected_category", "expected_subcategory",
            "expected_priority", "expected_sentiment",
            "predicted_category", "predicted_subcategory",
            "predicted_confidence", "predicted_priority", "predicted_sentiment",
            "category_correct",
            # dialog
            "conversation_turns", "conversation_text",
            "response", "context_gaps", "question_count",
            # general
            "test_type", "complexity", "prompt", "expected_behavior", "run_count",
            # rag
            "query", "context", "ground_truth",
            "groundedness_score", "relevance_score",
            # tool calling
            "available_tools", "expected_tool_calls",
            "tool_accuracy", "param_accuracy",
        ]
        _write_header(ws, columns)

        for item in raw:
            token_detail = item.get("token_detail") or {}
            # Handle list of token_details (general type has list)
            if isinstance(token_detail, list):
                td = token_detail[0] if token_detail else {}
            else:
                td = token_detail

            # --- Common ---
            scenario_id = (
                item.get("scenario_id")
                or item.get("test_id")
                or ""
            )
            latency = item.get("latency")
            tokens = item.get("tokens")
            prompt_tok = td.get("prompt", td.get("prompt_tokens"))
            completion_tok = td.get("completion", td.get("completion_tokens"))
            cached_tok = td.get("cached", td.get("cached_tokens"))
            reasoning_tok = td.get("reasoning", td.get("reasoning_tokens"))

            # --- Classification ---
            expected = item.get("expected") or {}
            predicted = item.get("predicted") or {}
            input_text = item.get("input", item.get("query"))
            expected_cat = expected.get("category")
            predicted_cat = predicted.get("category")
            category_correct = (
                (expected_cat == predicted_cat) if expected_cat and predicted_cat else None
            )

            # --- Dialog ---
            conversation = item.get("conversation")
            conversation_turns = len(conversation) if isinstance(conversation, list) else None
            conversation_text = None
            if isinstance(conversation, list):
                conversation_text = " | ".join(
                    f"{t.get('role', '?')}: {t.get('content', '')}"
                    for t in conversation
                )

            # --- General ---
            responses = item.get("responses")
            first_response = (
                responses[0] if isinstance(responses, list) and responses
                else item.get("response")
            )
            run_count = len(responses) if isinstance(responses, list) else None

            # --- RAG ---
            query = item.get("query")
            context = item.get("context")
            ground_truth = item.get("ground_truth")
            groundedness = item.get("groundedness")
            relevance = item.get("relevance")

            # --- Tool Calling ---
            available_tools = item.get("available_tools")
            expected_tools = item.get("expected_tool_calls")
            tool_acc = item.get("tool_accuracy")
            param_acc = item.get("param_accuracy")

            row = [
                model, eval_type, scenario_id,
                _safe_value(latency), _safe_value(tokens),
                _safe_value(prompt_tok), _safe_value(completion_tok),
                _safe_value(cached_tok), _safe_value(reasoning_tok),
                # classification
                _safe_value(input_text),
                expected_cat, expected.get("subcategory"),
                expected.get("priority"), expected.get("sentiment"),
                predicted_cat, predicted.get("subcategory"),
                _safe_value(predicted.get("confidence")),
                predicted.get("priority"), predicted.get("sentiment"),
                category_correct,
                # dialog
                conversation_turns, _safe_value(conversation_text),
                _safe_value(first_response),
                _safe_value(item.get("context_gaps")),
                item.get("question_count"),
                # general
                item.get("test_type"), item.get("complexity"),
                _safe_value(item.get("prompt")),
                _safe_value(item.get("expected_behavior")),
                run_count,
                # rag
                _safe_value(query), _safe_value(context), _safe_value(ground_truth),
                _safe_value(groundedness), _safe_value(relevance),
                # tool calling
                _safe_value(available_tools), _safe_value(expected_tools),
                _safe_value(tool_acc), _safe_value(param_acc),
            ]
            ws.append(row)

        _auto_width(ws)

    @staticmethod
    def _write_foundry_scores_eval(
        ws, foundry: Dict[str, Any], data: Dict[str, Any],
        model: str, eval_type: str,
    ) -> None:
        """Write FoundryScores sheet for an evaluation result."""
        _write_header(ws, [
            "model_name", "evaluation_type",
            "foundry_eval_id", "foundry_run_id", "foundry_report_url",
            "grader_name", "aggregated_score",
            "row_index", "row_score", "row_reason",
        ])
        eval_id = data.get("foundry_eval_id", "")
        run_id = data.get("foundry_run_id", "")
        report_url = data.get("foundry_report_url", "")
        aggregated = foundry.get("aggregated") or {}
        per_row = foundry.get("per_row") or []

        # Aggregated rows (row_index = -1)
        for grader, score in aggregated.items():
            ws.append([
                model, eval_type, eval_id, run_id, report_url,
                grader, _safe_value(score), -1, None, None,
            ])

        # Per-row detail
        for row_entry in per_row:
            row_idx = row_entry.get("row_index", 0)
            scores = row_entry.get("scores") or {}
            for grader, score_data in scores.items():
                if isinstance(score_data, dict):
                    ws.append([
                        model, eval_type, eval_id, run_id, report_url,
                        grader, None, row_idx,
                        _safe_value(score_data.get("score")),
                        score_data.get("reason", ""),
                    ])
                else:
                    ws.append([
                        model, eval_type, eval_id, run_id, report_url,
                        grader, None, row_idx, _safe_value(score_data), None,
                    ])
        _auto_width(ws)
